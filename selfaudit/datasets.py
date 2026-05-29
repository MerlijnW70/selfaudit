"""Dataset trust scanning: the engine for the fourth application of the harness.

This file contains *no* audit logic. It provides the interchangeable pieces the
controller in ``datasetscanner.py`` drives:

* :class:`Dataset`     - a loaded table (pure-stdlib CSV, or built in memory)
* :class:`CheckResult` - the raw outcome of one rule check
* :class:`Check`       - a named rule the scanner runs and re-tests
* the check builders   - ``values_in_range``, ``timestamps_monotonic``,
  ``no_missing_required``, ``duplicate_rate_below``, ``distribution_stationary``

Each check states a *hard, checkable invariant* over the data (the same shape as
``|f(x)| <= tol``). A failing check is the "unexpected" outcome the auditor then
re-tests by segment analysis to localize and classify the anomaly.
"""

from __future__ import annotations

import csv
import io
import json
import math
from collections.abc import Callable, Iterable
from dataclasses import dataclass, field
from datetime import date, datetime


@dataclass
class Dataset:
    """A loaded table: column names + rows as ``{column: value}`` dicts."""

    columns: list[str]
    rows: list[dict[str, str]]
    name: str = ""

    @property
    def n(self) -> int:
        return len(self.rows)

    def slice(self, start: int, end: int) -> Dataset:
        return Dataset(self.columns, self.rows[start:end], f"{self.name}[{start}:{end}]")

    def numeric_column(self, field_name: str) -> list[tuple[int, float | None]]:
        """Parse ``field_name`` per row to float; ``None`` for missing/unparseable."""
        out: list[tuple[int, float | None]] = []
        for i, row in enumerate(self.rows):
            raw = (row.get(field_name) or "").strip()
            if raw == "":
                out.append((i, None))
                continue
            try:
                out.append((i, float(raw)))
            except ValueError:
                out.append((i, None))
        return out


def _stringify(value: object) -> str:
    return "" if value is None else str(value)


def _decode(raw: bytes) -> str:
    """Decode bytes to text, tolerating a BOM and non-UTF-8 files (latin-1 last)."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _all_numeric(cells: list[str]) -> bool:
    """True if every non-empty cell parses as a number (used to spot a header-less
    first row of data)."""
    seen = False
    for c in cells:
        c = c.strip()
        if c == "":
            continue
        try:
            float(c)
            seen = True
        except ValueError:
            return False
    return seen


def parse_csv(text: str, name: str = "") -> Dataset:
    """Parse CSV/TSV *text* into a :class:`Dataset`.

    Auto-detects the delimiter (``,`` ``;`` tab ``|``) and whether a header row is
    present (a fully numeric first row is treated as data, with synthesized column
    names). Pure stdlib.
    """
    sample = text[:8192]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    raw_rows = [r for r in csv.reader(io.StringIO(text), delimiter=delimiter) if r]
    if not raw_rows:
        return Dataset([], [], name)
    if _all_numeric(raw_rows[0]):
        columns = [f"col{i + 1}" for i in range(len(raw_rows[0]))]
        data = raw_rows
    else:
        columns = [c.strip() for c in raw_rows[0]]
        data = raw_rows[1:]
    rows = [{columns[i]: (r[i] if i < len(r) else "") for i in range(len(columns))} for r in data]
    return Dataset(columns, rows, name)


def parse_json(text: str, name: str = "") -> Dataset:
    """Parse JSON *text* into a :class:`Dataset`.

    Accepts a list of objects (``[{"a":1}, ...]``) or column arrays
    (``{"a":[1,2], "b":[3,4]}`` — e.g. an Open-Meteo ``hourly`` block).
    """
    data = json.loads(text)
    if isinstance(data, list):
        objs = [o for o in data if isinstance(o, dict)]
        columns: list[str] = []
        for o in objs:
            for k in o:
                if k not in columns:
                    columns.append(k)
        rows = [{c: _stringify(o.get(c)) for c in columns} for o in objs]
        return Dataset(columns, rows, name)
    if isinstance(data, dict):
        arrays = {k: v for k, v in data.items() if isinstance(v, list)}
        if arrays:
            n = max(len(v) for v in arrays.values())
            cols = list(arrays)
            rows = [
                {c: (_stringify(arrays[c][i]) if i < len(arrays[c]) else "") for c in cols}
                for i in range(n)
            ]
            return Dataset(cols, rows, name)
    raise ValueError("unsupported JSON shape — need a list of objects or column arrays")


def parse_text(text: str, name: str = "") -> Dataset:
    """Parse text as JSON when the name ends in ``.json`` or the content looks like
    JSON, otherwise as CSV/TSV."""
    looks_json = name.lower().endswith(".json") or text.lstrip()[:1] in ("[", "{")
    if looks_json:
        try:
            return parse_json(text, name)
        except (ValueError, json.JSONDecodeError):
            if name.lower().endswith(".json"):
                raise
    return parse_csv(text, name)


def load_xlsx(path: str, *, name: str = "", sheet: str | None = None) -> Dataset:
    """Load the first row as headers and the rest as data from an ``.xlsx`` file.

    Requires the optional ``openpyxl`` package (``pip install selfaudit[excel]``).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without openpyxl
        raise ValueError(
            "Excel support needs the optional 'openpyxl' package (pip install 'selfaudit[excel]')"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    raw_rows = [list(r) for r in worksheet.iter_rows(values_only=True)]
    workbook.close()
    raw_rows = [r for r in raw_rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not raw_rows:
        return Dataset([], [], name or path)
    columns = [_stringify(c).strip() or f"col{i + 1}" for i, c in enumerate(raw_rows[0])]
    rows = [
        {columns[i]: _stringify(r[i]) if i < len(r) else "" for i in range(len(columns))}
        for r in raw_rows[1:]
    ]
    return Dataset(columns, rows, name or path)


def load_dataset(path: str, *, name: str = "") -> Dataset:
    """Load a dataset file by extension: ``.xlsx`` via openpyxl, ``.json`` as JSON,
    everything else as CSV/TSV (delimiter + encoding auto-detected)."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        return load_xlsx(path, name=name or path)
    with open(path, "rb") as fh:
        return parse_text(_decode(fh.read()), name or path)


def load_csv(path: str, *, name: str = "") -> Dataset:
    """Load a CSV/TSV file into a :class:`Dataset` (delimiter + encoding detected)."""
    with open(path, "rb") as fh:
        return parse_csv(_decode(fh.read()), name or path)


@dataclass
class CheckResult:
    """Raw outcome of one rule check (before the audit wraps it)."""

    ok: bool
    measured: float  # the measured quantity (a fraction or ratio; 0 == clean for most)
    threshold: float
    detail: str
    bad_rows: list[int] = field(default_factory=list)  # 0-based offending row indices


@dataclass
class Check:
    """A named rule: ``run(dataset) -> CheckResult``.

    ``severity`` rates how serious a violation is: ``"fail"`` (a hard problem ->
    UNTRUSTED), ``"warn"`` (worth review but often legitimate, e.g. outliers ->
    REVIEW), or ``"info"``. This is what keeps the scanner from crying wolf.
    """

    name: str
    run: Callable[[Dataset], CheckResult]
    severity: str = "fail"


def _fraction(count: int, total: int) -> float:
    return count / total if total else 0.0


def values_in_range(
    field_name: str, lo: float, hi: float, *, max_fraction: float = 0.0, severity: str = "fail"
) -> Check:
    """Numeric ``field_name`` must lie in ``[lo, hi]`` (missing values are ignored
    here — that is the missing-values check's job)."""

    def run(ds: Dataset) -> CheckResult:
        bad = [i for i, v in ds.numeric_column(field_name) if v is not None and not lo <= v <= hi]
        frac = _fraction(len(bad), ds.n)
        return CheckResult(
            ok=frac <= max_fraction,
            measured=frac,
            threshold=max_fraction,
            detail=f"{len(bad)}/{ds.n} rows have {field_name} outside [{lo:g}, {hi:g}]",
            bad_rows=bad,
        )

    return Check(f"range[{field_name}]", run, severity)


def timestamps_monotonic(
    field_name: str, *, max_fraction: float = 0.0, severity: str = "warn"
) -> Check:
    """Numeric ``field_name`` must be non-decreasing from row to row."""

    def run(ds: Dataset) -> CheckResult:
        values = ds.numeric_column(field_name)
        bad: list[int] = []
        prev: float | None = None
        for i, v in values:
            if v is not None and prev is not None and v < prev:
                bad.append(i)
            if v is not None:
                prev = v
        frac = _fraction(len(bad), ds.n)
        return CheckResult(
            ok=frac <= max_fraction,
            measured=frac,
            threshold=max_fraction,
            detail=f"{len(bad)}/{ds.n} rows where {field_name} decreases",
            bad_rows=bad,
        )

    return Check(f"monotonic[{field_name}]", run, severity)


def no_missing_required(
    fields: Iterable[str], *, max_fraction: float = 0.01, severity: str = "fail"
) -> Check:
    """Each required field must be missing in at most ``max_fraction`` of rows.

    Reported *per column* — the check fails if any single field exceeds the
    budget, and the detail names the offending columns with their individual
    missing rates (so "Cabin 77%, Age 20%" rather than a blunt aggregate).
    """
    required = list(fields)

    def run(ds: Dataset) -> CheckResult:
        fracs = {
            f: _fraction(sum(1 for row in ds.rows if (row.get(f) or "").strip() == ""), ds.n)
            for f in required
        }
        worst = max(fracs.values(), default=0.0)
        over = sorted((f for f in required if fracs[f] > max_fraction), key=lambda f: -fracs[f])
        bad = [
            i
            for i, row in enumerate(ds.rows)
            if any((row.get(f) or "").strip() == "" for f in over)
        ]
        if over:
            breakdown = ", ".join(f"{f} {fracs[f]:.1%}" for f in over)
            detail = f"columns over the {max_fraction:.1%} budget: {breakdown}"
        elif fracs:
            worst_field = max(fracs, key=lambda f: fracs[f])
            detail = f"all {len(required)} columns within budget (worst: {worst_field} {worst:.1%})"
        else:
            detail = "no columns to check"
        return CheckResult(
            ok=not over,
            measured=worst,
            threshold=max_fraction,
            detail=detail,
            bad_rows=bad,
        )

    return Check("missing_required", run, severity)


def duplicate_rate_below(*, max_fraction: float = 0.0, severity: str = "warn") -> Check:
    """At most ``max_fraction`` of rows may be exact duplicates of an earlier row."""

    def run(ds: Dataset) -> CheckResult:
        seen: set[tuple[tuple[str, str], ...]] = set()
        bad: list[int] = []
        for i, row in enumerate(ds.rows):
            key = tuple(sorted(row.items()))
            if key in seen:
                bad.append(i)
            else:
                seen.add(key)
        frac = _fraction(len(bad), ds.n)
        return CheckResult(
            ok=frac <= max_fraction,
            measured=frac,
            threshold=max_fraction,
            detail=f"{len(bad)}/{ds.n} duplicate rows ({frac:.1%})",
            bad_rows=bad,
        )

    return Check("duplicate_rate", run, severity)


def distribution_stationary(
    field_name: str, *, max_shift: float = 3.0, severity: str = "warn"
) -> Check:
    """The mean of ``field_name`` must not shift between the first and second half
    by more than ``max_shift`` pooled standard deviations (a regime-shift guard)."""

    def _stats(xs: list[float]) -> tuple[float, float]:
        if not xs:
            return 0.0, 0.0
        m = sum(xs) / len(xs)
        var = sum((x - m) ** 2 for x in xs) / len(xs)
        return m, math.sqrt(var)

    def run(ds: Dataset) -> CheckResult:
        present = [(i, v) for i, v in ds.numeric_column(field_name) if v is not None]
        if len(present) < 4:
            return CheckResult(True, 0.0, max_shift, f"too few {field_name} values to assess shift")
        mid = len(present) // 2
        first = [v for _, v in present[:mid]]
        second = [v for _, v in present[mid:]]
        m1, s1 = _stats(first)
        m2, s2 = _stats(second)
        pooled = math.sqrt((s1 * s1 + s2 * s2) / 2.0)
        shift = abs(m2 - m1) / pooled if pooled > 1e-12 else 0.0
        ok = shift <= max_shift
        bad = [i for i, _ in present[mid:]] if not ok else []
        return CheckResult(
            ok=ok,
            measured=shift,
            threshold=max_shift,
            detail=(
                f"{field_name} mean shifts {shift:.2f}σ between halves (1st={m1:.3g}, 2nd={m2:.3g})"
            ),
            bad_rows=bad,
        )

    return Check(f"stationary[{field_name}]", run, severity)


def unique_key(fields: Iterable[str], *, severity: str = "fail") -> Check:
    """The combination of ``fields`` must be unique across rows — a primary-key
    check (e.g. no duplicate ``customer_id``). Flags every row whose key repeats."""
    cols = list(fields)
    label = "+".join(cols)

    def run(ds: Dataset) -> CheckResult:
        seen: set[tuple[str, ...]] = set()
        bad: list[int] = []
        for i, row in enumerate(ds.rows):
            key = tuple((row.get(c) or "").strip() for c in cols)
            if key in seen:
                bad.append(i)
            else:
                seen.add(key)
        return CheckResult(
            ok=not bad,
            measured=_fraction(len(bad), ds.n),
            threshold=0.0,
            detail=f"{len(bad)}/{ds.n} rows duplicate the key ({label})",
            bad_rows=bad,
        )

    return Check(f"unique[{label}]", run, severity)


def _is_int(s: str) -> bool:
    try:
        int(s)
        return True
    except ValueError:
        return False


def _is_float(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False


def _is_bool(s: str) -> bool:
    return s.lower() in {"true", "false", "0", "1", "yes", "no", "y", "n", "t", "f"}


_DATE_FORMATS = ("%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S")


def _is_date(s: str) -> bool:
    for parser in (date.fromisoformat, datetime.fromisoformat):
        try:
            parser(s)
            return True
        except ValueError:
            pass
    for fmt in _DATE_FORMATS:
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            continue
    return False


_TYPE_PARSERS: dict[str, Callable[[str], bool]] = {
    "int": _is_int,
    "float": _is_float,
    "bool": _is_bool,
    "date": _is_date,
}


def values_of_type(
    field_name: str, type_name: str, *, max_fraction: float = 0.0, severity: str = "fail"
) -> Check:
    """Every non-empty value in ``field_name`` must parse as ``type_name``
    (``int`` / ``float`` / ``bool`` / ``date``). Missing values are ignored here."""
    parser = _TYPE_PARSERS[type_name]

    def run(ds: Dataset) -> CheckResult:
        bad = [
            i
            for i, row in enumerate(ds.rows)
            if (raw := (row.get(field_name) or "").strip()) != "" and not parser(raw)
        ]
        return CheckResult(
            ok=_fraction(len(bad), ds.n) <= max_fraction,
            measured=_fraction(len(bad), ds.n),
            threshold=max_fraction,
            detail=f"{len(bad)}/{ds.n} {field_name} values are not valid {type_name}",
            bad_rows=bad,
        )

    return Check(f"type[{field_name}={type_name}]", run, severity)


def allowed_values(field_name: str, allowed: Iterable[str], *, severity: str = "fail") -> Check:
    """Every non-empty value in ``field_name`` must be one of ``allowed`` (a
    categorical whitelist, e.g. ``status ∈ {active, churned}``)."""
    allowed_set = set(allowed)

    def run(ds: Dataset) -> CheckResult:
        bad = [
            i
            for i, row in enumerate(ds.rows)
            if (raw := (row.get(field_name) or "").strip()) != "" and raw not in allowed_set
        ]
        preview = ", ".join(sorted(allowed_set)[:6])
        return CheckResult(
            ok=not bad,
            measured=_fraction(len(bad), ds.n),
            threshold=0.0,
            detail=f"{len(bad)}/{ds.n} {field_name} values not in {{{preview}}}",
            bad_rows=bad,
        )

    return Check(f"allowed[{field_name}]", run, severity)


def _percentile(sorted_vals: list[float], q: float) -> float:
    """Linear-interpolated percentile of an already-sorted list (q in [0, 1])."""
    if not sorted_vals:
        return 0.0
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    pos = q * (len(sorted_vals) - 1)
    lo = int(pos)
    frac = pos - lo
    if lo + 1 < len(sorted_vals):
        return sorted_vals[lo] * (1 - frac) + sorted_vals[lo + 1] * frac
    return sorted_vals[lo]


def iqr_outliers(
    field_name: str, *, k: float = 3.0, max_fraction: float = 0.0, severity: str = "warn"
) -> Check:
    """Numeric ``field_name`` must have no values beyond Tukey fences
    ``[Q1 - k·IQR, Q3 + k·IQR]``.

    Robust and domain-free — the workhorse for vetting data whose valid range you
    do not know in advance (it learns the bulk and flags what falls far outside).
    """

    def run(ds: Dataset) -> CheckResult:
        present = [(i, v) for i, v in ds.numeric_column(field_name) if v is not None]
        if len(present) < 4:
            return CheckResult(
                True, 0.0, max_fraction, f"too few {field_name} values for an outlier test"
            )
        nums = sorted(v for _, v in present)
        q1, q3 = _percentile(nums, 0.25), _percentile(nums, 0.75)
        iqr = q3 - q1
        if iqr == 0.0:
            # No spread in the middle 50% (e.g. a zero-inflated count column):
            # IQR fences collapse to a point and would flag every deviation. The
            # outlier test is not meaningful here — report clean rather than noisy.
            return CheckResult(
                True, 0.0, max_fraction, f"{field_name} has zero IQR; outlier test not applicable"
            )
        lo, hi = q1 - k * iqr, q3 + k * iqr
        bad = [i for i, v in present if v < lo or v > hi]
        frac = _fraction(len(bad), ds.n)
        return CheckResult(
            ok=frac <= max_fraction,
            measured=frac,
            threshold=max_fraction,
            detail=f"{len(bad)}/{ds.n} {field_name} values beyond {k:g}·IQR fences "
            f"[{lo:.4g}, {hi:.4g}]",
            bad_rows=bad,
        )

    return Check(f"outliers[{field_name}]", run, severity)


def _looks_numeric(ds: Dataset, field_name: str) -> list[float]:
    """Return the parsed values if ``field_name`` is mostly numeric, else ``[]``."""
    col = ds.numeric_column(field_name)
    present = [v for _, v in col if v is not None]
    non_empty = sum(1 for row in ds.rows if (row.get(field_name) or "").strip() != "")
    if len(present) >= 8 and non_empty > 0 and len(present) >= 0.8 * non_empty:
        return present
    return []


def infer_checks(ds: Dataset, *, missing_fraction: float = 0.01) -> list[Check]:
    """Propose a sensible rule set from the data itself — zero configuration.

    Always checks missing-value budget and duplicate rows; for each numeric
    column with enough distinct values, adds an IQR-outlier and a regime-shift
    check, plus a monotonic check for columns that look like an ordered sequence
    (timestamps/ids). Categorical and low-cardinality columns are left alone.
    """
    checks: list[Check] = [
        no_missing_required(ds.columns, max_fraction=missing_fraction),
        duplicate_rate_below(max_fraction=0.0),
    ]
    for col in ds.columns:
        present = _looks_numeric(ds, col)
        if not present or len(set(present)) < 5:
            continue  # non-numeric, too sparse, or low-cardinality (codes/flags)
        decreases = sum(1 for a, b in zip(present, present[1:], strict=False) if b < a)
        is_ordered = decreases <= 0.05 * len(present) and present[-1] > present[0]
        checks.append(iqr_outliers(col, k=3.0))
        if is_ordered:
            # An ordered sequence (timestamp/index): a monotonic check is meaningful,
            # but a stationarity check is not — a trending mean is expected, not an anomaly.
            checks.append(timestamps_monotonic(col))
        else:
            checks.append(distribution_stationary(col, max_shift=3.0))
    return checks


def _is_ordered(vals: list[float]) -> bool:
    decreases = sum(1 for a, b in zip(vals, vals[1:], strict=False) if b < a)
    return decreases <= 0.05 * len(vals) and vals[-1] > vals[0]


def svg_chart(ds: Dataset, *, width: int = 860, height: int = 220, max_series: int = 3) -> str:
    """An inline SVG line chart of the dataset's *value* columns over row order.

    Index/timestamp-like columns (monotonic sequences) are skipped — they are the
    x-axis, not data. Each series is normalized to its own range so columns of
    different scales fit one frame. Returns ``""`` when there is nothing to plot.
    Pure stdlib; the SVG is self-contained and escapes all labels.
    """
    from html import escape

    series: list[tuple[str, list[float]]] = []
    for col in ds.columns:
        vals = [v for _, v in ds.numeric_column(col) if v is not None]
        if len(vals) < 3 or len(set(vals)) < 3 or _is_ordered(vals):
            continue
        series.append((col, vals))
        if len(series) >= max_series:
            break
    if not series:
        return ""

    pad = 30
    iw, ih = width - 2 * pad, height - 2 * pad
    colors = ["#1f883d", "#cf222e", "#9a6700"]
    parts = [
        f"<svg viewBox='0 0 {width} {height}' width='100%' role='img' "
        f"style='max-height:{height}px'>",
        f"<rect x='{pad}' y='{pad}' width='{iw}' height='{ih}' fill='#fff' stroke='#e1e6ea'/>",
    ]
    for g in (0.0, 0.5, 1.0):
        y = pad + ih * g
        parts.append(
            f"<line x1='{pad}' y1='{y:.1f}' x2='{pad + iw}' y2='{y:.1f}' stroke='#eef1f4'/>"
        )
    legend = []
    for idx, (col, vals) in enumerate(series):
        lo, hi = min(vals), max(vals)
        rng = (hi - lo) or 1.0
        n = len(vals)
        pts = " ".join(
            f"{pad + iw * i / (n - 1):.1f},{pad + ih * (1 - (v - lo) / rng):.1f}"
            for i, v in enumerate(vals)
        )
        color = colors[idx % len(colors)]
        parts.append(f"<polyline fill='none' stroke='{color}' stroke-width='2' points='{pts}'/>")
        legend.append(
            f"<span style='color:{color}'>●</span> {escape(col)} "
            f"<span class='muted'>[{lo:.4g} – {hi:.4g}, n={n}]</span>"
        )
    parts.append("</svg>")
    return "".join(parts) + f"<div class='legend'>{' &nbsp; '.join(legend)}</div>"
